"""Convert XLSX to markdown tables."""
import pathlib
from openpyxl import load_workbook


def xlsx_to_markdown(pad: pathlib.Path) -> str:
    wb = load_workbook(pad, read_only=True, data_only=True)
    parts = []

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        # Filter out completely empty rows
        rows = [r for r in rows if any(c is not None for c in r)]
        if not rows:
            continue

        parts.append(f"### {sheet.title}")

        header = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join(["---"] * len(header)) + " |")

        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            # Pad or trim to match header width
            while len(cells) < len(header):
                cells.append("")
            parts.append("| " + " | ".join(cells[:len(header)]) + " |")

    return "\n\n".join(parts)
